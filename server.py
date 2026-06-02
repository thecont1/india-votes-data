"""
ECI Results Scraper - Server Entry Point (FastAPI)

This module provides the FastAPI server for ECI results scraping
and the live election dashboard API.
"""

import os
import re
import sys
import time
from datetime import datetime
from typing import Optional

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from io import BytesIO
from pydantic import BaseModel

from db import _connect, _cursor, IS_PG, get_elections, get_current_election, get_election_by_id

app = FastAPI(
    title="ECI Results Scraper API",
    description="API for scraping Election Commission of India election results",
    version="2.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# Party colors for dashboard (custom palette, not ECI raw hex)
# ---------------------------------------------------------------------------
PARTY_COLORS = {
    "BJP": "#FF6B00",
    "INC": "#00BFFF",
    "DMK": "#FF0000",
    "AIADMK": "#008000",
    "AITC": "#00FF7F",
    "CPM": "#CC0000",
    "TVK": "#FFD700",
    "INC": "#1E90FF",
    "IUML": "#006400",
    "AINRC": "#808080",
    "CPI": "#8B0000",
    "CPI(M)": "#CC0000",
    "BPF": "#00CED1",
    "AGP": "#32CD32",
    "VCK": "#8A2BE2",
    "PMK": "#A9A9A9",
    "IND": "#D3D3D3",
    "NCP": "#00008B",
    "JD(U)": "#008080",
    "SHS": "#FF4500",
    "TDP": "#FFD700",
    "YSRCP": "#1E90FF",
    "AAP": "#0066CC",
    "BRS": "#FF69B4",
}
DEFAULT_COLOR = "#888888"

# Lazy-loaded party symbol URLs (abv -> url)
_party_symbols: Optional[dict] = None


def _get_party_symbols() -> dict:
    global _party_symbols
    if _party_symbols is None:
        try:
            conn = _connect()
            cur = _cursor(conn)
            cur.execute("SELECT abv, symbol_url FROM parties WHERE symbol_url IS NOT NULL")
            _party_symbols = {row["abv"]: row["symbol_url"] for row in cur.fetchall()}
            conn.close()
        except Exception:
            _party_symbols = {}
    return _party_symbols


# ---------------------------------------------------------------------------
# Pydantic models (scraping)
# ---------------------------------------------------------------------------
class ScrapeRequest(BaseModel):
    url: str
    limit: int = 3
    respect: bool = False


class ScrapeAcRoundsRequest(BaseModel):
    url: str
    ac_no: int
    start_round: int = 1


class ScrapeAllRoundsRequest(BaseModel):
    url: str
    start_ac: int = 1
    end_ac: int = 0
    respect: bool = False


# ---------------------------------------------------------------------------
# Dashboard API
# ---------------------------------------------------------------------------
@app.get("/api/seat-tally")
def seat_tally(
    state: str = Query("", description="State code filter, empty=all"),
    election_id: str = Query("", description="Election ID to filter states"),
):
    """Party-wise won + leading seat counts plus deposit-loss breakdown.

    Returns list of {party_abv, party_name, color, won, leading, total,
                      lost_no_deposit, lost_deposit} sorted by won descending.
    """
    conn = _connect()
    cur = _cursor(conn)
    try:
        sf = ""                               # state filter fragment
        params: list = []
        if election_id:
            # Get states for this election and filter by them
            election = get_election_by_id(election_id)
            if election:
                state_list = election["states"]
                if len(state_list) == 1:
                    sf = "AND r.state_code = {}".format("%s" if IS_PG else "?")
                    params.append(state_list[0])
                    params.append(state_list[0])  # appears in 2 CTEs
                else:
                    sf = "AND r.state_code IN ({})".format(
                        ",".join(["%s"] * len(state_list)) if IS_PG else ",".join(["?"] * len(state_list))
                    )
                    params.extend(state_list)
                    params.extend(state_list)  # appears in 2 CTEs
        elif state:
            sf = "AND r.state_code = {}".format("%s" if IS_PG else "?")
            params.append(state)
            params.append(state)  # appears in 2 CTEs

        query = f"""
        WITH latest_rounds AS (
            SELECT lr.state_code, lr.ac_no, lr.max_round
            FROM (
                SELECT state_code, ac_no, MAX(round_no) as max_round
                FROM rounds_ac
                GROUP BY state_code, ac_no
            ) lr
            JOIN constituency_status cs
                ON lr.state_code = cs.state_code AND lr.ac_no = cs.ac_no
            WHERE cs.status = 'DONE'
        ),
        ac_totals AS (
            SELECT r.state_code, r.ac_no, SUM(r.votes) as total_votes
            FROM rounds_ac r
            JOIN latest_rounds lr
                ON r.state_code = lr.state_code AND r.ac_no = lr.ac_no
                AND r.round_no = lr.max_round
            GROUP BY r.state_code, r.ac_no
        ),
        ac_winners AS (
            SELECT state_code, ac_no, winner_abv
            FROM (
                SELECT lr.state_code, lr.ac_no, p.abv as winner_abv,
                       ROW_NUMBER() OVER (
                           PARTITION BY lr.state_code, lr.ac_no
                           ORDER BY r.votes DESC
                       ) as rn
                FROM rounds_ac r
                JOIN latest_rounds lr
                    ON r.state_code = lr.state_code AND r.ac_no = lr.ac_no
                    AND r.round_no = lr.max_round
                JOIN parties p ON r.party_abv = p.abv
                WHERE 1=1 {sf}
            ) WHERE rn = 1
        ),
        party_best AS (
            SELECT state_code, ac_no, party_abv, party_name,
                   votes, ac_declared, winner_abv, total_votes
            FROM (
                SELECT r.state_code, r.ac_no,
                       p.abv as party_abv, p.name as party_name,
                       r.votes, cs.won as ac_declared,
                       aw.winner_abv, at.total_votes,
                       ROW_NUMBER() OVER (
                           PARTITION BY r.state_code, r.ac_no, p.abv
                           ORDER BY r.votes DESC
                       ) as rn
                FROM rounds_ac r
                JOIN latest_rounds lr
                    ON r.state_code = lr.state_code AND r.ac_no = lr.ac_no
                    AND r.round_no = lr.max_round
                JOIN constituency_status cs
                    ON r.state_code = cs.state_code AND r.ac_no = cs.ac_no
                JOIN parties p ON r.party_abv = p.abv
                JOIN ac_totals at
                    ON r.state_code = at.state_code AND r.ac_no = at.ac_no
                JOIN ac_winners aw
                    ON r.state_code = aw.state_code AND r.ac_no = aw.ac_no
                WHERE 1=1 {sf}
            ) WHERE rn = 1
        )
        SELECT
            party_abv,
            MAX(party_name) as party_name,
            SUM(CASE WHEN party_abv = winner_abv AND ac_declared = 1
                     THEN 1 ELSE 0 END) as won_seats,
            SUM(CASE WHEN party_abv = winner_abv AND ac_declared = 0
                     THEN 1 ELSE 0 END) as leading_seats,
            SUM(CASE WHEN party_abv != winner_abv AND ac_declared = 1
                     AND votes * 6 >= total_votes
                     THEN 1 ELSE 0 END) as lost_no_deposit,
            SUM(CASE WHEN party_abv != winner_abv AND ac_declared = 1
                     AND votes * 6 < total_votes
                     THEN 1 ELSE 0 END) as lost_deposit
            , SUM(votes) as total_votes
        FROM party_best
        GROUP BY party_abv
        ORDER BY won_seats DESC
        """
        cur.execute(query, params)
        rows = cur.fetchall()

        # Check if won status is populated at all (historical data may have won=0 everywhere)
        check_q = "SELECT SUM(CASE WHEN won=1 THEN 1 ELSE 0 END) as won_count FROM constituency_status"
        if state:
            p = "%s" if IS_PG else "?"
            check_q += f" WHERE state_code={p}"
        cur.execute(check_q, [state] if state else [])
        has_won_data = (cur.fetchone() or {}).get("won_count", 0) > 0

        result = []
        symbols = _get_party_symbols()
        for row in rows:
            abv = row["party_abv"]
            won = row["won_seats"]
            leading = row["leading_seats"]
            lost_no_dep = row["lost_no_deposit"]
            lost_dep = row["lost_deposit"]
            # If no won status populated anywhere, treat all as won (historical data)
            if not has_won_data:
                won += leading
                leading = 0
                lost_no_dep = 0
                lost_dep = 0
            result.append({
                "party_abv": abv,
                "party_name": row.get("party_name", abv),
                "won": won,
                "leading": leading,
                "total": won + leading,
                "lost_no_deposit": lost_no_dep,
                "lost_deposit": lost_dep,
                "total_votes": row.get("total_votes", 0),
                "color": PARTY_COLORS.get(abv, DEFAULT_COLOR),
                "symbol_url": symbols.get(abv),
            })

        # Compute majority line from states with DONE ACs
        majority = None
        if state:
            cur.execute(
                f"SELECT assembly_seats FROM states WHERE state_code={('%s' if IS_PG else '?')}",
                (state,),
            )
            row = cur.fetchone()
            if row:
                majority = row["assembly_seats"] // 2 + 1
        else:
            # Overall: sum assembly_seats only for states that have DONE ACs
            cur.execute(
                "SELECT SUM(s.assembly_seats) as total_seats "
                "FROM states s "
                "JOIN (SELECT DISTINCT state_code FROM constituency_status WHERE status = 'DONE') cs "
                "ON s.state_code = cs.state_code"
            )
            row = cur.fetchone()
            if row and row["total_seats"]:
                majority = row["total_seats"] // 2 + 1

        return {
            "parties": result,
            "majority": majority,
            "updated_at": datetime.now().isoformat(),
        }
    finally:
        conn.close()


@app.get("/api/parties")
def get_parties():
    """All party details for legend tooltips."""
    conn = _connect()
    cur = _cursor(conn)
    try:
        cur.execute("""
            SELECT abv, name, chief, founded,
                   seats_loksabha, seats_rajyasabha, seats_assembly,
                   wikipedia_url, alliance, symbol_url
            FROM parties
            ORDER BY abv
        """)
        rows = cur.fetchall()
        result = {}
        for row in rows:
            d = dict(row) if hasattr(row, 'keys') else {
                'abv': row[0], 'name': row[1], 'chief': row[2],
                'founded': row[3], 'seats_loksabha': row[4],
                'seats_rajyasabha': row[5], 'seats_assembly': row[6],
                'wikipedia_url': row[7], 'alliance': row[8],
                'symbol_url': row[9],
            }
            result[d['abv']] = d
        return {"parties": result}
    finally:
        conn.close()


@app.get("/api/ac-races")
def ac_races(state: str = Query(..., description="State code (required)")):
    """Per-AC candidate data: all candidates in each AC's latest round.

    Returns every AC in the state with all candidates ranked by votes.
    Status is computed from actual rounds_ac data (not constituency_status).
    """
    conn = _connect()
    cur = _cursor(conn)
    try:
        p = "%s" if IS_PG else "?"
        cur.execute(f"""
            WITH latest_rounds AS (
                SELECT lr.state_code, lr.ac_no, lr.max_round
                FROM (
                    SELECT state_code, ac_no, MAX(round_no) as max_round
                    FROM rounds_ac
                    WHERE state_code = {p}
                    GROUP BY state_code, ac_no
                ) lr
                JOIN constituency_status cs
                    ON lr.state_code = cs.state_code AND lr.ac_no = cs.ac_no
                WHERE cs.status = 'DONE'
            ),
            ranked AS (
                SELECT r.state_code, r.ac_no, r.ac_name,
                       r.candidate,
                       p.abv as party_abv, p.name as party_name,
                       r.votes,
                       cs.current_round,
                       cs.won,
                       lr.max_round as latest_round,
                       cs.form20_url,
                       cs.form20_status,
                       cs.form20_score,
                       cs.form20_checked_at,
                       ROW_NUMBER() OVER (
                           PARTITION BY r.state_code, r.ac_no
                           ORDER BY r.votes DESC
                       ) as rank
                FROM rounds_ac r
                JOIN latest_rounds lr
                    ON r.state_code = lr.state_code
                    AND r.ac_no = lr.ac_no
                    AND r.round_no = lr.max_round
                JOIN constituency_status cs
                    ON r.state_code = cs.state_code
                    AND r.ac_no = cs.ac_no
                JOIN parties p ON r.party_abv = p.abv
            )
            SELECT ac_no, ac_name, candidate, party_abv, party_name,
                   votes, rank,
                   current_round, won, latest_round,
                   form20_url, form20_status,
                   form20_score, form20_checked_at,
                   SUM(votes) OVER (PARTITION BY ac_no) as total_votes
            FROM ranked
            ORDER BY ac_no, rank
        """, (state,))
        rows = cur.fetchall()

        # Group by AC
        from collections import OrderedDict
        ac_map = OrderedDict()
        symbols = _get_party_symbols()
        for row in rows:
            d = dict(row) if hasattr(row, 'keys') else {
                'ac_no': row[0], 'ac_name': row[1], 'candidate': row[2],
                'party_abv': row[3], 'party_name': row[4],
                'votes': row[5], 'rank': row[6],
                'current_round': row[7], 'won': row[8], 'latest_round': row[9],
                'form20_url': row[10], 'form20_status': row[11],
                'form20_score': row[12], 'form20_checked_at': row[13],
                'total_votes': row[14],
            }
            ac_no = d['ac_no']
            if ac_no not in ac_map:
                ac_map[ac_no] = {
                    'ac_no': ac_no,
                    'ac_name': d['ac_name'],
                    'total_votes': d['total_votes'],
                    'status': 'PENDING',  # Will be computed after all candidates collected
                    'current_round': d.get('current_round', 0),
                    'won': d.get('won', 0),
                    'latest_round': d['latest_round'],
                    'form20_url': d.get('form20_url'),
                    'form20_status': d.get('form20_status', 'UNAVAILABLE'),
                    'form20_score': d.get('form20_score'),
                    'form20_checked_at': d.get('form20_checked_at'),
                    'margin': 0,
                    'candidates': [],
                }
            d['color'] = PARTY_COLORS.get(d['party_abv'], DEFAULT_COLOR)
            d['symbol_url'] = symbols.get(d['party_abv'])
            ac_map[ac_no]['candidates'].append(d)

        # Compute status for each AC based on actual data
        for ac in ac_map.values():
            cands = ac['candidates']
            total_votes = ac['total_votes']
            n_candidates = len(cands)
            # DONE criteria: votes > 0 and >1 candidate (counting rounds, not 999)
            if total_votes > 0 and n_candidates > 1:
                ac['status'] = 'DONE'

        result = list(ac_map.values())
        # Set margin = winner votes - runner-up votes
        for ac in result:
            cands = ac['candidates']
            if len(cands) >= 2:
                ac['margin'] = cands[0]['votes'] - cands[1]['votes']
            elif len(cands) == 1:
                ac['margin'] = cands[0]['votes']
        # Sort by margin descending (largest margin first = default selection)
        result.sort(key=lambda a: a['margin'], reverse=True)

        return {"races": result, "state": state}
    finally:
        conn.close()


@app.get("/api/roundwise")
def roundwise(state: str = Query(..., description="State code (required)")):
    """Roundwise progression: cumulative votes as ACs complete counting.

    For each target round R and each AC, pick the effective round as the
    latest available round snapshot where round_no <= R.  The leader's
    vote count at that effective round is the AC's contribution.

    Two phases:
    - Rounds 1..N: per-AC latest-snapshot leaders (excluding round 999)
    - Round F (999): final tally including postal ballots (may flip some
      winners)
    """
    from bisect import bisect_right

    conn = _connect()
    cur = _cursor(conn)
    try:
        p = "%s" if IS_PG else "?"

        # Phase 1: counting rounds (exclude 999)
        # For each AC and each round, find the party leading at that round.
        cur.execute(f"""
            WITH ranked AS (
                SELECT state_code, ac_no, round_no, party_abv, votes,
                       ROW_NUMBER() OVER (
                           PARTITION BY state_code, ac_no, round_no
                           ORDER BY votes DESC
                       ) as rn
                FROM rounds_ac
                WHERE state_code = {p} AND round_no != 999
            )
            SELECT ac_no, round_no, party_abv, votes
            FROM ranked WHERE rn = 1
            ORDER BY ac_no, round_no
        """, (state,))
        rows = cur.fetchall()

        from collections import defaultdict
        all_parties = set()

        # Build per-AC sorted round data for binary-search lookup
        ac_data = {}          # ac_no -> [(round_no, party_abv, votes)]
        ac_round_keys = {}    # ac_no -> [round_no] (sorted, for bisect)
        for row in rows:
            rd = dict(row) if hasattr(row, 'keys') else {
                'ac_no': row[0], 'round_no': row[1],
                'party_abv': row[2], 'votes': row[3],
            }
            ac_no = rd['ac_no']
            if ac_no not in ac_data:
                ac_data[ac_no] = []
                ac_round_keys[ac_no] = []
            ac_data[ac_no].append((rd['round_no'], rd['party_abv'], rd['votes']))
            ac_round_keys[ac_no].append(rd['round_no'])
            all_parties.add(rd['party_abv'])

        # Distinct target rounds from actual data (excluding 999)
        distinct_rounds = sorted(set(
            dict(r)['round_no'] if hasattr(r, 'keys') else r[1]
            for r in rows
        ))

        # For each target round, compute party totals using effective rounds.
        # For each AC, binary-search the latest round <= target_rn.
        counting_data = defaultdict(lambda: defaultdict(int))
        for target_rn in distinct_rounds:
            for ac_no in ac_data:
                rnds = ac_round_keys[ac_no]
                idx = bisect_right(rnds, target_rn) - 1
                if idx >= 0:
                    _, party, votes = ac_data[ac_no][idx]
                    counting_data[target_rn][party] += votes

        # Phase 2: F round (999) — final winners including postal ballots
        cur.execute(f"""
            WITH ranked AS (
                SELECT r.state_code, r.ac_no,
                       p.abv as party_abv, r.votes,
                       ROW_NUMBER() OVER (
                           PARTITION BY r.state_code, r.ac_no
                           ORDER BY r.votes DESC
                       ) as rank
                FROM rounds_ac r
                JOIN parties p ON r.party_abv = p.abv
                WHERE r.state_code = {p} AND r.round_no = 999
            )
            SELECT party_abv, SUM(votes) as total_votes
            FROM ranked
            WHERE rank = 1
            GROUP BY party_abv
        """, (state,))
        f_rows = cur.fetchall()
        f_data = defaultdict(int)
        for row in f_rows:
            rd = dict(row) if hasattr(row, 'keys') else {
                'party_abv': row[0], 'total_votes': row[1],
            }
            f_data[rd['party_abv']] = rd['total_votes']
            all_parties.add(rd['party_abv'])

        # Build all_rounds from actual distinct round numbers present
        all_rounds = distinct_rounds[:]
        has_f = len(f_data) > 0
        if has_f:
            all_rounds.append(999)

        # Each round's totals are already complete snapshots (not incremental)
        cumulative_series = {}
        for rn in all_rounds:
            if rn == 999:
                cumulative_series[rn] = dict(f_data) if has_f else {}
            else:
                cumulative_series[rn] = dict(counting_data.get(rn, {}))

        # Sort parties by final vote count
        final_key = 999 if has_f else all_rounds[-1] if all_rounds else None
        final_votes = {p: cumulative_series.get(final_key, {}).get(p, 0)
                       for p in all_parties} if final_key is not None else {}
        sorted_parties = sorted(all_parties, key=lambda p: final_votes.get(p, 0),
                                reverse=True)

        series = []
        symbols = _get_party_symbols()
        for party in sorted_parties:
            if final_votes.get(party, 0) == 0:
                continue
            series.append({
                'party_abv': party,
                'party_name': party,
                'color': PARTY_COLORS.get(party, DEFAULT_COLOR),
                'symbol_url': symbols.get(party),
                'data': [cumulative_series[rn].get(party, 0) for rn in all_rounds],
            })

        return {
            "state": state,
            "rounds": all_rounds,
            "series": series,
        }
    finally:
        conn.close()


@app.get("/api/status")
def status_summary(
    state: str = Query(default=None),
    election_id: str = Query(default=None),
):
    """Counting progress summary — reads directly from constituency_status.

    Each writer (insert_round_snapshot, upsert_constituency_status)
    keeps constituency_status in sync, so it's the source of truth.
    NULL status is treated as PENDING (results not yet scraped).
    """
    conn = _connect()
    cur = _cursor(conn)
    try:
        p = "%s" if IS_PG else "?"
        state_filter = ""
        state_params: list = []
        if state:
            state_filter = f"AND cs.state_code = {p}"
            state_params.append(state)
        elif election_id:
            election = get_election_by_id(election_id)
            if election:
                state_list = election["states"]
                if len(state_list) == 1:
                    state_filter = f"AND cs.state_code = {p}"
                    state_params.append(state_list[0])
                else:
                    state_filter = "AND cs.state_code IN ({})".format(
                        ",".join(["%s"] * len(state_list)) if IS_PG
                        else ",".join(["?"] * len(state_list))
                    )
                    state_params.extend(state_list)

        cur.execute(f"""
            SELECT
                COALESCE(cs.status, 'PENDING') as effective_status,
                COUNT(*) as cnt
            FROM constituency_status cs
            WHERE cs.state_code IN (SELECT DISTINCT state_code FROM rounds_ac)
              {state_filter}
            GROUP BY effective_status
        """, state_params)
        statuses = {row["effective_status"]: row["cnt"] for row in cur.fetchall()}
        return {
            "statuses": statuses,
            "active_states": 1 if state else len(statuses),
            "updated_at": datetime.now().isoformat(),
        }
    finally:
        conn.close()


@app.get("/api/constituency/{state_code}/{ac_no}")
def constituency_rounds(state_code: str, ac_no: int):
    """Round-by-round data for one AC."""
    conn = _connect()
    cur = _cursor(conn)
    try:
        cur.execute(
            f"SELECT round_no, candidate, party_abv, votes "
            f"FROM rounds_ac WHERE state_code={('%s' if IS_PG else '?')} "
            f"AND ac_no={('%s' if IS_PG else '?')} ORDER BY round_no, votes DESC",
            (state_code, ac_no),
        )
        rows = cur.fetchall()
        return {"state_code": state_code, "ac_no": ac_no, "rounds": rows}
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Election management endpoints
# ---------------------------------------------------------------------------
@app.get("/api/elections")
def list_elections():
    """List all elections ordered by sort_date descending."""
    return {"elections": get_elections()}


@app.get("/api/elections/current")
def current_election():
    """Get the current (most recent) election."""
    election = get_current_election()
    if not election:
        return {"election": None}
    return {"election": election}


@app.get("/api/states")
def list_states():
    """Return all states with their codes and names."""
    conn = _connect()
    cur = _cursor(conn)
    try:
        cur.execute("SELECT state_code, state_name FROM states ORDER BY state_code")
        return {"states": cur.fetchall()}
    finally:
        conn.close()


@app.get("/api/download")
def download_dataset(
    state: str = Query(default=None, description="State code filter"),
    election_id: str = Query(default=None, description="Election ID filter"),
):
    """Return a branded Excel (.xlsx) file with joined EVM + postal vote data.

    Sheet 1 ("Data"):  per-candidate EVM and postal votes, styled to match
                       the LET Live dashboard palette.
    Sheet 2 ("About"): source / website / author / provenance metadata.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Brand palette — light theme for print-friendly spreadsheets
    C_BG     = "FFFFFF"   # white background
    C_HEADER = "FFFFFF"   # white header
    C_CARD   = "FFFFFF"   # data rows (odd)
    C_CARD2  = "F8F9FA"   # zebra stripe (even)
    C_TEXT   = "1A1A1A"   # near-black text
    C_MUTED  = "6C757D"   # grey muted
    C_GOLD   = "B8860B"   # dark gold for light bg
    C_BORDER = "DEE2E6"   # light border

    conn = _connect()
    cur  = _cursor(conn)
    try:
        p = "%s" if IS_PG else "?"

        state_codes: list[str] = []
        state_names: dict[str, str] = {}
        election_name = ""
        if state:
            state_codes = [state]
        if election_id:
            election = get_election_by_id(election_id)
            if election:
                election_name = election["name"]
                if not state:
                    state_codes = election["states"]

        if state_codes:
            q = (
                f"SELECT state_code, state_name, state_code_std FROM states "
                f"WHERE state_code IN ({','.join([p]*len(state_codes))})"
            )
            for row in cur.execute(q, state_codes).fetchall():
                state_names[row["state_code"]] = row["state_name"]

        sf_plain = ""
        sf       = ""
        params: list = []
        if state_codes:
            sf_plain = f"AND state_code IN ({','.join([p]*len(state_codes))})"
            sf       = f"AND r.state_code IN ({','.join([p]*len(state_codes))})"
            params   = list(state_codes)

        cur.execute(f"""
            WITH latest_non999 AS (
                SELECT state_code, ac_no, MAX(round_no) as max_round
                FROM rounds_ac
                WHERE round_no != 999 {sf_plain}
                GROUP BY state_code, ac_no
            ),
            all_candidates AS (
                SELECT DISTINCT r.state_code, r.ac_no, r.ac_name, r.candidate, r.party_abv
                FROM rounds_ac r
                JOIN latest_non999 lr
                  ON r.state_code = lr.state_code AND r.ac_no = lr.ac_no AND r.round_no = lr.max_round
                UNION
                SELECT DISTINCT r.state_code, r.ac_no, r.ac_name, r.candidate, r.party_abv
                FROM rounds_ac r
                WHERE r.round_no = 999 {sf}
            ),
            evm AS (
                SELECT r.state_code, r.ac_no, r.candidate, r.party_abv, r.votes as evm_votes
                FROM rounds_ac r
                JOIN latest_non999 lr
                  ON r.state_code = lr.state_code AND r.ac_no = lr.ac_no AND r.round_no = lr.max_round
            ),
            total AS (
                SELECT r.state_code, r.ac_no, r.candidate, r.party_abv, r.votes as total_votes
                FROM rounds_ac r
                WHERE r.round_no = 999 {sf}
            )
            SELECT
                ac.state_code, st.state_name, ac.ac_no, ac.ac_name, ac.candidate,
                p.abv  as party_abv,
                p.name as party_name,
                COALESCE(e.evm_votes, 0)                                as evm_votes,
                COALESCE(t.total_votes, 0) - COALESCE(e.evm_votes, 0) as postal_votes,
                COALESCE(t.total_votes, 0)                             as total_votes
            FROM all_candidates ac
            LEFT JOIN states st ON ac.state_code = st.state_code
            JOIN parties p ON ac.party_abv = p.abv
            LEFT JOIN evm e
              ON ac.state_code = e.state_code AND ac.ac_no = e.ac_no
             AND ac.candidate  = e.candidate  AND ac.party_abv = e.party_abv
            LEFT JOIN total t
              ON ac.state_code = t.state_code AND ac.ac_no = t.ac_no
             AND ac.candidate  = t.candidate  AND ac.party_abv = t.party_abv
            ORDER BY ac.state_code, ac.ac_no,
                (COALESCE(e.evm_votes,0) + COALESCE(t.total_votes,0) - COALESCE(e.evm_votes,0)) DESC
        """, params + params + params)
        rows = cur.fetchall()
    finally:
        conn.close()

    # Normalise to dicts
    row_data: list[dict] = []
    for r in rows:
        d = dict(r) if hasattr(r, "keys") else {
            "state_code": r[0], "state_name": r[1], "ac_no": r[2],
            "ac_name": r[3], "candidate": r[4], "party_abv": r[5],
            "party_name": r[6], "evm_votes": r[7],
            "postal_votes": r[8], "total_votes": r[9],
        }
        row_data.append(d)

    # ── Workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_GOLD

    headers = [
        "State Code", "State Name", "AC No.", "AC Name", "Candidate",
        "Party Abbr.", "Party Name",
        "EVM Votes", "Postal Votes", "Total Votes",
    ]
    num_cols = len(headers)

    # Row 1 — title banner with logo
    ws.row_dimensions[1].height = 42

    # Logo image (favicon)
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'favicon.png')
    if os.path.exists(logo_path):
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo_path)
            img.width = 32
            img.height = 32
            ws.add_image(img, 'A1')
        except Exception:
            pass  # logo not available, text-only header is fine

    # Title text (merged B1:H1, leaving A1 for the logo)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=num_cols)
    t = ws.cell(row=1, column=2)
    t.value     = "LET Live!  —  The Live & Loaded Elections Tracker of India"
    t.font      = Font(name="Calibri", bold=True, size=18, color=C_GOLD)
    t.fill      = PatternFill("solid", fgColor=C_HEADER)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=C_HEADER)

    # Row 2 — subtitle / datestamp
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    sub_parts = []
    if election_name:
        sub_parts.append(election_name)
    if len(state_names) == 1:
        sub_parts.append(list(state_names.values())[0])
    elif len(state_names) > 1:
        sub_parts.append(", ".join(state_names.values()))
    sub_parts.append(f"Downloaded {datetime.now().strftime('%d %b %Y, %H:%M')} IST")
    sub_parts.append("results.eci.gov.in")
    s = ws.cell(row=2, column=1)
    s.value     = "  ·  ".join(sub_parts)
    s.font      = Font(name="Calibri", italic=True, size=9, color=C_MUTED)
    s.fill      = PatternFill("solid", fgColor=C_BG)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Row 3 — spacer
    for col in range(1, num_cols + 1):
        ws.cell(row=3, column=col).fill = PatternFill("solid", fgColor=C_BG)
    ws.row_dimensions[3].height = 5

    # Row 4 — column headers
    HEADER_ROW  = 4
    gold_border = Border(bottom=Side(style="medium", color=C_GOLD))
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=h.upper())
        cell.font      = Font(name="Calibri", bold=True, size=9, color=C_TEXT)
        cell.fill      = PatternFill("solid", fgColor=C_CARD)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = gold_border
    ws.row_dimensions[HEADER_ROW].height = 20
    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    # Rows 5+ — data
    subtle_border = Border(bottom=Side(style="hair", color=C_BORDER))
    for row_idx, d in enumerate(row_data):
        excel_row  = HEADER_ROW + 1 + row_idx
        bg         = C_CARD if row_idx % 2 == 0 else C_CARD2
        party_hex  = PARTY_COLORS.get(d["party_abv"], DEFAULT_COLOR).lstrip("#")
        party_left = Border(
            left=Side(style="medium", color=party_hex),
            bottom=Side(style="hair", color=C_BORDER),
        )
        values = [
            d["state_code"], d["state_name"], d["ac_no"], d["ac_name"],
            d["candidate"], d["party_abv"], d["party_name"],
            d["evm_votes"], d["postal_votes"], d["total_votes"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell        = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.font   = Font(name="Calibri", size=9, color=C_TEXT)
            cell.border = party_left if col_idx == 1 else subtle_border
            if col_idx >= 8:
                cell.alignment     = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            elif col_idx in (1, 3):
                cell.alignment = Alignment(horizontal="center")

    # Column widths
    sample_end = min(len(row_data) + HEADER_ROW + 1, HEADER_ROW + 501)
    for col_idx, h in enumerate(headers, 1):
        max_len = len(h)
        for r in range(HEADER_ROW, sample_end):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows       = f"1:{HEADER_ROW}"
    ws.oddHeader.center.text  = "LET Live — India Elections Data"
    ws.oddFooter.left.text    = "Source: results.eci.gov.in"
    ws.oddFooter.center.text  = "&P of &N"
    ws.oddFooter.right.text   = "&D"

    # ── About sheet ─────────────────────────────────────────────────────
    about = wb.create_sheet("About")
    about.sheet_view.showGridLines   = False
    about.sheet_properties.tabColor  = "888888"
    about.column_dimensions["A"].width = 20
    about.column_dimensions["B"].width = 72

    about.merge_cells("A1:B1")
    ah = about["A1"]
    ah.value     = "LET Live!  —  Dataset Provenance"
    ah.font      = Font(name="Calibri", bold=True, size=13, color=C_GOLD)
    ah.fill      = PatternFill("solid", fgColor=C_HEADER)
    ah.alignment = Alignment(horizontal="left", vertical="center")
    about.row_dimensions[1].height = 28

    meta = [
        ("Source",         "Election Commission of India (ECI)"),
        ("Website",        "https://results.eci.gov.in"),
        ("Author",         "Mahesh Shantaram"),
        ("Author Email",   "ms@thecontrarian.in"),
        ("Author Website", "https://thecontrarian.in/"),
        ("Generated",      datetime.now().strftime("%Y-%m-%d %H:%M:%S IST")),
        ("Rows",           str(len(row_data))),
    ]
    if election_name:
        meta.append(("Election", election_name))
    if len(state_names) == 1:
        meta.append(("State", list(state_names.values())[0]))
    elif len(state_names) > 1:
        meta.append(("States", ", ".join(state_names.values())))

    for i, (key, val) in enumerate(meta, 2):
        bg  = C_BG if i % 2 == 0 else C_CARD2
        kc  = about.cell(row=i, column=1, value=key)
        kc.font      = Font(name="Calibri", bold=True, size=9, color=C_GOLD)
        kc.fill      = PatternFill("solid", fgColor=bg)
        kc.alignment = Alignment(horizontal="left")
        vc           = about.cell(row=i, column=2, value=val)
        vc.font      = Font(name="Calibri", size=9, color=C_TEXT)
        vc.fill      = PatternFill("solid", fgColor=bg)
        vc.alignment = Alignment(horizontal="left")

    # ── Serialise & stream ──────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Build filename: LETLive_<Election>_<StateCode><StateName>.xlsx
    name_parts = ["LETLive"]
    if election_name:
        name_parts.append(election_name.replace(" ", "").replace("-", ""))
    if len(state_codes) == 1:
        code = state_codes[0]
        state_name_clean = state_names.get(code, "").replace(" ", "")
        name_parts.append(f"{code}{state_name_clean}")
    filename = "_".join(name_parts) + ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Scraping endpoints
# ---------------------------------------------------------------------------
@app.post("/scrape")
async def scrape_endpoint(request: ScrapeRequest):
    """Scrape constituency results from ECI party-wise URL."""
    from core.browser import create_chrome_driver
    from core.scraper import (
        build_constituency_url, get_state_code,
        parse_partywise_url, scrape_constituency_sync,
    )
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait

    try:
        election_identifier, state_code = parse_partywise_url(request.url)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    results = scrape_constituency_sync(
        election_identifier, state_code,
        limit=request.limit,
        respect_mode=request.respect
    )

    if results["constituencywise_results"]:
        driver = create_chrome_driver()
        try:
            url = build_constituency_url(election_identifier, state_code, 1)
            driver.get(url)
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, 'h1'))
            )
            h1 = driver.find_element(By.TAG_NAME, 'h1').text
            h2 = driver.find_element(By.TAG_NAME, 'h2').text
            state_name = h2.split('(')[-1].replace(')', '')
            results["election_year"] = h1.split('-')[-1].strip()
            results["election_type"] = ''.join(h2.split()[:1])
            results["election_state"] = get_state_code(state_name)
        except Exception:
            pass
        finally:
            driver.quit()

    return {"status": "success", "data": results}


@app.get("/health")
async def health_check():
    return {"status": "healthy"}


@app.post("/scrape/ac-rounds")
def scrape_ac_rounds_endpoint(request: ScrapeAcRoundsRequest):
    from core.browser import create_chrome_driver
    from core.scraper import parse_partywise_url, scrape_ac_rounds_core

    try:
        election_identifier, state_code = parse_partywise_url(request.url)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    driver = create_chrome_driver()
    try:
        result = scrape_ac_rounds_core(
            driver, election_identifier, state_code,
            request.ac_no, request.start_round
        )
        if result.get("status") == "done":
            return {"status": "error", "error": "AC not found (404)"}
        return result
    except Exception as e:
        return {"status": "error", "error": str(e)}
    finally:
        driver.quit()


@app.post("/scrape/all-rounds")
def scrape_all_rounds_endpoint(request: ScrapeAllRoundsRequest):
    from core.browser import create_chrome_driver
    from core.scraper import build_constituency_url, parse_partywise_url, scrape_ac_rounds_core

    try:
        election_identifier, state_code = parse_partywise_url(request.url)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    driver = create_chrome_driver()
    results = []

    try:
        if request.end_ac == 0:
            max_ac = 0
            for i in range(1, 1000):
                test_url = build_constituency_url(election_identifier, state_code, i)
                driver.get(test_url)
                if "404" in driver.title:
                    break
                max_ac = i
        else:
            max_ac = request.end_ac

        for ac_no in range(request.start_ac, max_ac + 1):
            result = scrape_ac_rounds_core(
                driver, election_identifier, state_code, ac_no, 1
            )
            if result.get("status") == "done":
                break
            if result.get("status") == "success":
                results.append(result.get("data", {}))
            if request.respect and ac_no % 10 == 0:
                time.sleep(1)

    except Exception as e:
        return {"status": "error", "error": str(e)}
    finally:
        driver.quit()

    return {"status": "success", "data": results, "total_acs": len(results)}



import csv

@app.get("/api/tv-channels")
def tv_channels():
    """Return TV channel list from data/tv.csv."""
    tv_csv = os.path.join(os.path.dirname(__file__), "data", "tv.csv")
    if not os.path.exists(tv_csv):
        return {"channels": []}
    channels = []
    with open(tv_csv, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            link = row.get("link", "")
            # Extract YouTube video_id from watch URL
            m = re.search(r"v=([A-Za-z0-9_-]+)", link)
            if m:
                channels.append({
                    "name": row.get("channel", ""),
                    "language": row.get("language", ""),
                    "video_id": m.group(1),
                    "link": link,
                })
    return {"channels": channels}

# ---------------------------------------------------------------------------
# Static file serving (dashboard must be mounted LAST)
# ---------------------------------------------------------------------------
STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")

if os.path.isdir(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.get("/")
def serve_dashboard():
    """Serve the live dashboard."""
    index = os.path.join(STATIC_DIR, "index.html")
    if os.path.exists(index):
        return FileResponse(index)
    return {"message": "Dashboard not found. Place index.html in static/"}


@app.get("/favicon.png")
def serve_favicon_png():
    f = os.path.join(STATIC_DIR, "favicon.png")
    if os.path.exists(f):
        return FileResponse(f, media_type="image/png")


@app.get("/favicon.ico")
def serve_favicon_ico():
    f = os.path.join(STATIC_DIR, "favicon.ico")
    if os.path.exists(f):
        return FileResponse(f, media_type="image/x-icon")


if __name__ == "__main__":
    if "--api" in sys.argv:
        import uvicorn
        uvicorn.run(app, host="0.0.0.0", port=8000)
    else:
        print("Use cli.py for command-line scraping, or run with --api flag.")
